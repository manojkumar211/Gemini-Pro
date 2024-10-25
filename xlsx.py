import json
import streamlit as st
import os
import google.generativeai as genai
from dotenv import load_dotenv
from openpyxl import load_workbook
from datetime import datetime, time


load_dotenv()


genai.configure(api_key=os.getenv("GOOGLE_API_KEY"))  


model = genai.GenerativeModel("gemini-pro")

st.set_page_config(page_title="Requirements")
st.header("Gemini LLM Model")


input_text = st.text_input("Input:", key="input")
upload_file = st.file_uploader("Upload a file", type='xlsx')
submit = st.button("Submit Here")

def serialize_value(value):
    if isinstance(value, (datetime, time)):
        return value.isoformat()  
    elif isinstance(value, (str, int, float)):
        return value  
    return str(value)  

def get_gen_response(input_text, headers, contents):
    
    if input_text:
        combined_input = f"{input_text}\nHeaders: {headers}\nContents: {contents}"
        response = model.generate_content([combined_input])
        return response.content  
    return None

def process_uploaded_file(upload_file):
    if upload_file.name.endswith('.xlsx'):
        workbook = load_workbook(upload_file)
        sheet = workbook.active
        structured_data = []

        
        headers = [cell.value for cell in sheet[1]]  
        
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            for i, cell in enumerate(row):
                if cell: 
                    structured_data.append({
                        "header": headers[i],
                        "content": serialize_value(cell)  
                    })

        
        json_file_path = "C:/Projects/astria/files/xlsx_new_json.json"  
        with open(json_file_path, 'w', encoding='utf-8') as json_file:
            json.dump(structured_data, json_file, indent=4)

        return structured_data  
    else:
        st.error("The uploaded file is not an XLSX file.")
        return None

if submit:
    structured_data = process_uploaded_file(upload_file)
    if structured_data:
        headers = [item['header'] for item in structured_data]
        contents = [item['content'] for item in structured_data]
        response = get_gen_response(input_text, headers, contents)
        if response:
            st.subheader("The Final Answer")
            st.write(response)
